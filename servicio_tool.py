import json
import sys
from pathlib import Path
from collections import defaultdict
from rapidfuzz import process

from openpyxl import Workbook, load_workbook

# When running as compiled exe (PyInstaller), use the exe's folder so config/data live next to it
BASE_DIR = Path(sys.executable).parent if getattr(sys, "frozen", False) else Path.cwd()
CONFIG_PATH = BASE_DIR / "config.json"
GROUPS_EXCEL = BASE_DIR / "groups.xlsx"
CALIS_EXCEL = BASE_DIR / "calis.xlsx"
GROUPS_DIR = BASE_DIR / "groups"


DEFAULT_CONFIG = {
    "MATCH_THRESHOLD": 90,
    "INCLUDE_MATCH_SCORE": False,
    "GROUPS_HEADER_ROW": 3,
    "CALIS_HEADER_ROW": 2,
    "COL_ACTIVITY": 8,
    "COL_NAME_GROUPS": 4,
    "COL_VALUE_GROUPS": 13,
    "COL_NAME_CALIS": 3,
}


def load_config():
    if not CONFIG_PATH.exists():
        with open(CONFIG_PATH, "w", encoding="utf-8") as f:
            json.dump(DEFAULT_CONFIG, f, indent=2)
        print(f"Created {CONFIG_PATH} with default config.")
    data = DEFAULT_CONFIG.copy()
    if CONFIG_PATH.exists():
        with open(CONFIG_PATH, encoding="utf-8") as f:
            data.update(json.load(f))
    return data


def discover_materias(base_dir):
    """
    Discover materias and grupos from folder structure: groups/<materia>/<grupo>.xlsx
    Returns e.g. {"SO": ["001", "002"], "BD": ["010"]}
    """
    if not base_dir.is_dir():
        return {}
    out = {}
    for materia_dir in sorted(base_dir.iterdir()):
        if not materia_dir.is_dir():
            continue
        grupos = [
            p.stem for p in sorted(materia_dir.iterdir())
            if p.suffix.lower() == ".xlsx"
        ]
        if grupos:
            out[materia_dir.name] = grupos
    return out


def find_match(query, choices, threshold):
    """Return (matched_string, score) or None if no match above threshold."""
    if not choices:
        return None
    best_match = process.extractOne(query, choices)
    if best_match[1] <= threshold:
        return None
    return best_match


def normalize_name(name):
    if name is None or (isinstance(name, str) and not name.strip()):
        return ""
    s = str(name).upper().removeprefix("ITS ").removeprefix("IAS ")
    parts = s.split()
    parts.sort()
    return " ".join(parts)


def normalize_names(names):
    return [normalize_name(n) for n in names]


def main():
    cfg = load_config()
    materias = discover_materias(GROUPS_DIR)
    if not materias:
        print(f"No materias/grupos found under {GROUPS_DIR} (expected structure: groups/<materia>/<grupo>.xlsx)")
        return

    result_wb = Workbook()
    # Remove default sheet so we only have our result sheets
    result_wb.remove(result_wb.active)

    if not CALIS_EXCEL.exists():
        print(f"ERROR: {CALIS_EXCEL} not found. Exiting.")
        return
    wb_calis = load_workbook(CALIS_EXCEL, read_only=False)

    GROUPS_HEADER_ROW = cfg["GROUPS_HEADER_ROW"]
    CALIS_HEADER_ROW = cfg["CALIS_HEADER_ROW"]
    # Convert 1-based column numbers to 0-based for indexing row tuples
    COL_ACTIVITY = cfg["COL_ACTIVITY"] - 1
    COL_NAME_GROUPS = cfg["COL_NAME_GROUPS"] - 1
    COL_VALUE_GROUPS = cfg["COL_VALUE_GROUPS"] - 1
    COL_NAME_CALIS = cfg["COL_NAME_CALIS"] - 1
    MATCH_THRESHOLD = cfg["MATCH_THRESHOLD"]
    INCLUDE_MATCH_SCORE = cfg["INCLUDE_MATCH_SCORE"]

    for materia, grupos in materias.items():
        for grupo in grupos:
            excel_path = GROUPS_DIR / materia / f"{grupo}.xlsx"
            if not excel_path.exists():
                print(f"Skip: {excel_path} not found")
                continue

            wb = load_workbook(excel_path)
            first_sheet = wb[wb.sheetnames[0]]
            assignments = defaultdict(list)

            for row in first_sheet.iter_rows(min_row=GROUPS_HEADER_ROW, values_only=True):
                if row is None or len(row) <= max(COL_ACTIVITY, COL_NAME_GROUPS, COL_VALUE_GROUPS):
                    continue
                act, name, val = row[COL_ACTIVITY], row[COL_NAME_GROUPS], row[COL_VALUE_GROUPS]
                if act is None:
                    continue
                assignments[act].append((normalize_name(name), val))

            all_names = {pair[0] for data in assignments.values() for pair in data if pair[0]}
            if not all_names:
                print(f"Skip: no names in {excel_path}")
                continue

            sheet_name = f"{grupo} {materia}"
            if sheet_name not in wb_calis.sheetnames:
                print(f"Skip: sheet '{sheet_name}' not in calis.xlsx")
                continue

            sheet = wb_calis[sheet_name]
            cal_names = []
            for row in sheet.iter_rows(min_row=CALIS_HEADER_ROW, values_only=True):
                if row is None or not any(row):
                    continue
                if len(row) > COL_NAME_CALIS and row[COL_NAME_CALIS] is not None:
                    cal_names.append(row[COL_NAME_CALIS])
            new_names = normalize_names(cal_names)

            header = ["name", "teams_name"]
            if INCLUDE_MATCH_SCORE:
                header.append("match_score")
            header.extend(assignments.keys())
            results = [header]

            for original, new_name in zip(cal_names, new_names):
                match = find_match(new_name, all_names, MATCH_THRESHOLD)
                if match:
                    matched_name, score = match[0], match[1]
                else:
                    matched_name, score = "NO MATCH", 0
                res = [original, matched_name]
                if INCLUDE_MATCH_SCORE:
                    res.append(score)
                for data in assignments.values():
                    mapped_cal = dict(data)
                    res.append(mapped_cal.get(matched_name, 0))
                results.append(res)

            result_ws = result_wb.create_sheet(title=f"{grupo} {materia}")
            for row in results:
                result_ws.append(row)

    result_wb.save(GROUPS_EXCEL)


if __name__ == "__main__":
    print(f"Archivos esperados en: {BASE_DIR}")
    main()
    x = input("press any  key to close the program")